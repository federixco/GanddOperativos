import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
import os
import io
import matplotlib.pyplot as plt
import matplotlib.patches as patches
from datetime import datetime

def exportar_a_excel(procesos_data, algoritmo, gantt_data, metricas, trm, tem, quantum=None):
    """
    Exporta los resultados a un archivo Excel con gráfico de Gantt PNG.
    
    Args:
        procesos_data: Lista de procesos originales
        algoritmo: Nombre del algoritmo usado
        gantt_data: Datos del gráfico de Gantt
        metricas: Lista de métricas calculadas
        trm: Tiempo de respuesta medio
        tem: Tiempo de espera medio
        quantum: Quantum usado (opcional)
    """
    
    # Crear workbook
    wb = Workbook()
    
    # Eliminar hoja por defecto
    wb.remove(wb.active)
    
    # Crear hojas
    ws_metricas = wb.create_sheet("Métricas")
    ws_gantt = wb.create_sheet("Gráfico Gantt")
    ws_datos = wb.create_sheet("Datos Originales")
    
    # === HOJA DE MÉTRICAS ===
    _crear_hoja_metricas(ws_metricas, metricas, trm, tem, algoritmo, quantum)
    
    # === HOJA DE GRÁFICO GANTT ===
    _crear_hoja_gantt_con_imagen(ws_gantt, gantt_data, algoritmo)
    
    # === HOJA DE DATOS ORIGINALES ===
    _crear_hoja_datos_originales(ws_datos, procesos_data)
    
    # Generar nombre de archivo
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"resultados_{algoritmo}_{timestamp}.xlsx"
    
    # Guardar archivo
    filepath = os.path.join(os.path.dirname(__file__), filename)
    wb.save(filepath)
    
    return filepath

def _crear_hoja_metricas(ws, metricas, trm, tem, algoritmo, quantum):
    """Crea la hoja de métricas."""
    ws.title = "Métricas"
    
    # Título
    ws['A1'] = f"Resultados - Algoritmo {algoritmo}"
    ws['A1'].font = Font(size=16, bold=True)
    
    if quantum:
        ws['A2'] = f"Quantum: {quantum}"
        ws['A2'].font = Font(size=12, italic=True)
    
    # Métricas generales
    ws['A4'] = "MÉTRICAS GENERALES"
    ws['A4'].font = Font(size=14, bold=True)
    ws['A4'].fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
    
    ws['A5'] = "TRM (Tiempo de Respuesta Medio):"
    ws['B5'] = trm
    ws['A6'] = "TEM (Tiempo de Espera Medio):"
    ws['B6'] = tem
    
    # Tabla de métricas por proceso
    ws['A8'] = "MÉTRICAS POR PROCESO"
    ws['A8'].font = Font(size=14, bold=True)
    ws['A8'].fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
    
    # Encabezados
    headers = ["PID", "Tiempo de Llegada", "Tiempo de CPU", "Tiempo de Respuesta", "Tiempo de Espera"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=9, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Datos
    for row, metrica in enumerate(metricas, 10):
        ws.cell(row=row, column=1, value=metrica["PID"])
        ws.cell(row=row, column=2, value=metrica["Llegada"])
        ws.cell(row=row, column=3, value=metrica["CPU"])
        ws.cell(row=row, column=4, value=metrica["TR"])
        ws.cell(row=row, column=5, value=metrica["TE"])
        
        # Formatear números
        for col in [2, 3, 4, 5]:
            ws.cell(row=row, column=col).number_format = "0.00"
    
    # Ajustar ancho de columnas
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws.column_dimensions[col].width = 15

def _crear_hoja_gantt_con_imagen(ws, gantt_data, algoritmo):
    """Crea la hoja con gráfico de Gantt como imagen PNG."""
    ws.title = "Gráfico Gantt"
    
    # Procesar datos de Gantt
    procesos_unicos = []
    gantt_procesado = []
    
    for segmento in gantt_data:
        if len(segmento) == 3:
            pid, start, end = segmento
            tipo = "CPU" if pid != "IDLE" else "IDLE"
        elif len(segmento) == 4:
            pid, start, end, tipo = segmento
        else:
            continue
            
        if pid != "IDLE" and pid not in procesos_unicos:
            procesos_unicos.append(pid)
            
        gantt_procesado.append({
            "PID": pid,
            "Inicio": start,
            "Fin": end,
            "Duración": end - start,
            "Tipo": tipo
        })
    
    # Crear DataFrame
    df_gantt = pd.DataFrame(gantt_procesado)
    
    # Escribir datos
    for r in dataframe_to_rows(df_gantt, index=False, header=True):
        ws.append(r)
    
    # Formatear encabezados
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Generar gráfico de Gantt como imagen
    _generar_gantt_png(ws, gantt_procesado, algoritmo)

def _generar_gantt_png(ws, gantt_procesado, algoritmo):
    """Genera un gráfico de Gantt como imagen PNG replicando exactamente la lógica del programa original."""
    if not gantt_procesado:
        return
    
    # Normalizar datos igual que en el programa original
    norm = []
    for segmento in gantt_procesado:
        pid = segmento['PID']
        start = segmento['Inicio']
        end = segmento['Fin']
        tipo = segmento['Tipo']
        
        # Convertir tipo a formato del programa original
        if tipo == "Bloqueo":
            tipo = "BLOCK"
        elif tipo == "CPU":
            tipo = "CPU"
        elif pid == "IDLE":
            tipo = "IDLE"
            
        norm.append((pid, start, end, tipo))
    
    # Obtener procesos únicos (excluir IDLE) - igual que el programa original
    procesos_unicos = [pid for pid, _, _, tipo in norm if pid != "IDLE"]
    procesos_unicos = list(dict.fromkeys(procesos_unicos))  # Mantener orden
    
    if not procesos_unicos:
        return
    
    # Calcular dimensiones adaptativas - igual que el programa original
    num_procesos = len(procesos_unicos)
    max_time = max(end for _, _, end, _ in norm) if norm else 1
    
    fig_width = max(15, max_time * 0.25)  # Ancho basado en duración total
    fig_height = max(4, num_procesos * 0.8)  # Altura basada en número de procesos
    
    # Crear figura con las mismas dimensiones
    fig, ax = plt.subplots(figsize=(fig_width, fig_height))
    
    # Posiciones Y para cada proceso - igual que el programa original
    y_positions = {pid: i for i, pid in enumerate(procesos_unicos)}
    
    # Colores - usar la misma paleta que el programa original
    colors = {}
    color_palette = plt.cm.get_cmap("tab20", len(procesos_unicos) + 1)
    
    # Dibujar barras - replicar exactamente la lógica del programa original
    for pid, start, end, tipo in norm:
        if tipo == "IDLE":
            color = "lightgray"
            hatch = None
            y = -1
            alpha = 0.7
        elif tipo == "BLOCK":
            color = "darkred"
            hatch = "///"
            y = y_positions.get(pid, 0)
            alpha = 0.8
        else:  # CPU
            if pid not in colors:
                colors[pid] = color_palette(len(colors))
            color = colors[pid]
            hatch = None
            y = y_positions.get(pid, 0)
            alpha = 1.0
        
        # Dibujar barra horizontal - mismo estilo que el programa original
        ax.barh(y, end - start, left=start, height=0.6, color=color, 
               edgecolor='black', hatch=hatch, alpha=alpha)
        
        if tipo != "IDLE":
            # Mostrar el PID y el tipo de ráfaga - mismo formato que el programa original
            text = f"{pid}\n({tipo})" if tipo == "BLOCK" else str(pid)
            ax.text((start + end) / 2, y, text, ha='center', va='center',
                    fontsize=7, color="white" if tipo == "BLOCK" else "black",
                    weight="bold" if tipo == "BLOCK" else "normal")
    
    # Configurar ejes - igual que el programa original
    if procesos_unicos:
        ax.set_yticks(list(y_positions.values()))
        ax.set_yticklabels(list(y_positions.keys()))
    else:
        ax.set_yticks([])
    
    if norm:
        max_time = max(end for _, _, end, _ in norm)
        # Mostrar TODOS los ticks del 0 al tiempo máximo - igual que el programa original
        ax.set_xticks(range(0, max_time + 1))
        ax.set_xlim(0, max_time)
    
    ax.set_xlabel("Tiempo")
    ax.set_ylabel("Procesos")
    ax.set_title(f"Diagrama de Gantt - {algoritmo}")
    ax.grid(True, axis='x', linestyle='--', alpha=0.6)
    
    # Leyenda explicativa - igual que el programa original
    legend_elements = [
        plt.Rectangle((0,0),1,1, facecolor='lightblue', edgecolor='black', label='CPU'),
        plt.Rectangle((0,0),1,1, facecolor='darkred', edgecolor='black', hatch='///', label='Bloqueo (E/S)'),
        plt.Rectangle((0,0),1,1, facecolor='lightgray', edgecolor='black', alpha=0.7, label='IDLE')
    ]
    ax.legend(handles=legend_elements, loc='upper left', bbox_to_anchor=(1.02, 1))
    
    # Ajustar layout
    plt.tight_layout()
    
    # Guardar imagen en memoria
    img_buffer = io.BytesIO()
    plt.savefig(img_buffer, format='png', dpi=300, bbox_inches='tight')
    img_buffer.seek(0)
    
    # Insertar imagen en Excel
    img = Image(img_buffer)
    img.width = 800  # Ancho más grande para mejor visualización
    img.height = 400  # Alto más grande
    
    # Posicionar imagen debajo de los datos
    ws.add_image(img, 'A8')
    
    # Cerrar figura para liberar memoria
    plt.close(fig)

def _crear_hoja_datos_originales(ws, procesos_data):
    """Crea la hoja con los datos originales de entrada."""
    ws.title = "Datos Originales"
    
    # Título
    ws['A1'] = "Datos Originales de los Procesos"
    ws['A1'].font = Font(size=16, bold=True)
    
    # Encabezados
    headers = ["PID", "Tiempo de Llegada", "Secuencia de Ráfagas"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Datos
    for row, proceso in enumerate(procesos_data, 4):
        ws.cell(row=row, column=1, value=proceso["pid"])
        ws.cell(row=row, column=2, value=proceso["arrival_time"])
        
        # Formatear ráfagas como texto
        bursts_text = " -> ".join([f"{'CPU' if i%2==0 else 'E/S'}: {burst}" 
                                  for i, burst in enumerate(proceso["bursts"])])
        ws.cell(row=row, column=3, value=bursts_text)
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 50
